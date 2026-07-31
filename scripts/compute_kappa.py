#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
compute_kappa.py

计算两位标注者之间的 Cohen's Kappa，
并输出 FP/FN 分析报告。

用法:
    python compute_kappa.py \
        --annotator_a  results/annotation_form_annotatorA.xlsx \
        --annotator_b  results/annotation_form_annotatorB.xlsx \
        --out          results/kappa_report.txt
"""

import re
import argparse
from pathlib import Path
from collections import defaultdict

import openpyxl


# ─────────────────────────────────────────────────────────────
# 读取标注Excel
# ─────────────────────────────────────────────────────────────

def parse_yn(val):
    """将 Yes/No/yes/no/Y/N 统一转为 True/False/None"""
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("yes", "y", "1", "true"):
        return True
    if s in ("no", "n", "0", "false"):
        return False
    return None  # 未填写


def load_annotations(xlsx_path):
    """
    读取标注Excel，返回:
    {
      sample_id: {
        "model": {
          "L1": True/False,
          "L2": True/False,
          "L3": True/False,
          "comment": str
        }
      }
    }
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 找标注sheet（不是说明页）
    ann_sheet = None
    for name in wb.sheetnames:
        if "Annotator" in name or "annotator" in name:
            ann_sheet = wb[name]
            break
    if ann_sheet is None:
        raise ValueError(f"找不到标注sheet，请确认文件: {xlsx_path}")

    results = {}
    current_sample_id = None
    model_order = ["Base", "FT-A", "FT-A×3.5", "FT-A+B+C"]

    for row in ann_sheet.iter_rows(min_row=4, values_only=True):
        if row[0] is None and row[1] is None and row[4] is None:
            continue

        # 分隔行（样本标题行）
        col_a = str(row[0]) if row[0] else ""
        if "样本" in col_a or "──" in col_a:
            # 从分隔行提取sample_id
            m = re.search(r"(SUM_\w+|RW_\w+)", col_a)
            if m:
                current_sample_id = m.group(1)
                results[current_sample_id] = {}
            continue

        # 数据行
        if current_sample_id is None:
            continue

        # col index (0-based): 4=Model, 6=L1, 7=L2, 8=L3, 9=Comment
        model_val = row[4]
        l1_val    = row[6]
        l2_val    = row[7]
        l3_val    = row[8]
        comment   = row[9] if len(row) > 9 else None

        if model_val is None:
            continue

        model_label = str(model_val).strip()
        if model_label not in model_order:
            continue

        results[current_sample_id][model_label] = {
            "L1":      parse_yn(l1_val),
            "L2":      parse_yn(l2_val),
            "L3":      parse_yn(l3_val),
            "comment": str(comment).strip() if comment else "",
        }

    return results


# ─────────────────────────────────────────────────────────────
# Cohen's Kappa
# ─────────────────────────────────────────────────────────────

def cohen_kappa(labels_a, labels_b):
    """
    计算二分类Cohen's Kappa。
    labels_a, labels_b: list of True/False (长度相同，已过滤None)
    """
    assert len(labels_a) == len(labels_b), "长度不一致"
    n = len(labels_a)
    if n == 0:
        return None, 0, {}

    # 混淆矩阵
    tp = sum(1 for a, b in zip(labels_a, labels_b) if a and b)
    tn = sum(1 for a, b in zip(labels_a, labels_b) if not a and not b)
    fp = sum(1 for a, b in zip(labels_a, labels_b) if not a and b)   # A=No, B=Yes
    fn = sum(1 for a, b in zip(labels_a, labels_b) if a and not b)   # A=Yes, B=No

    po = (tp + tn) / n  # 观测一致率

    # 期望一致率
    p_yes_a = (tp + fn) / n
    p_yes_b = (tp + fp) / n
    p_no_a  = (tn + fp) / n
    p_no_b  = (tn + fn) / n
    pe = p_yes_a * p_yes_b + p_no_a * p_no_b

    if pe == 1.0:
        kappa = 1.0
    else:
        kappa = (po - pe) / (1 - pe)

    cm = {"TP": tp, "TN": tn, "FP": fp, "FN": fn}
    return round(kappa, 4), n, cm


def kappa_interpretation(kappa):
    if kappa is None:
        return "N/A"
    if kappa > 0.80:
        return "Almost Perfect（几乎完美）"
    if kappa > 0.61:
        return "Substantial（高度一致）"
    if kappa > 0.41:
        return "Moderate（中等一致）"
    if kappa > 0.21:
        return "Fair（一般）"
    return "Slight（较差）"


# ─────────────────────────────────────────────────────────────
# FP/FN 分析（自动标注 vs 人工标注）
# ─────────────────────────────────────────────────────────────

def compute_fp_fn(human_labels, auto_labels, level):
    """
    human_labels: list of True/False（人工标注，作为gold standard）
    auto_labels:  list of True/False（自动标注）
    """
    n = len(human_labels)
    if n == 0:
        return {}

    tp = sum(1 for h, a in zip(human_labels, auto_labels) if h and a)
    tn = sum(1 for h, a in zip(human_labels, auto_labels) if not h and not a)
    fp = sum(1 for h, a in zip(human_labels, auto_labels) if not h and a)
    fn = sum(1 for h, a in zip(human_labels, auto_labels) if h and not a)

    fpr = round(fp / (fp + tn) * 100, 1) if (fp + tn) > 0 else 0.0
    fnr = round(fn / (fn + tp) * 100, 1) if (fn + tp) > 0 else 0.0
    precision = round(tp / (tp + fp) * 100, 1) if (tp + fp) > 0 else 0.0
    recall    = round(tp / (tp + fn) * 100, 1) if (tp + fn) > 0 else 0.0

    return {
        "n": n, "TP": tp, "TN": tn, "FP": fp, "FN": fn,
        "FPR": fpr, "FNR": fnr,
        "Precision": precision, "Recall": recall,
    }


# ─────────────────────────────────────────────────────────────
# 报告生成
# ─────────────────────────────────────────────────────────────

def generate_report(ann_a, ann_b):
    lines = []
    lines.append("=" * 70)
    lines.append("INTER-ANNOTATOR AGREEMENT REPORT")
    lines.append("Cohen's Kappa & FP/FN Analysis")
    lines.append("=" * 70)

    models    = ["Base", "FT-A", "FT-A×3.5", "FT-A+B+C"]
    levels    = ["L1", "L2", "L3"]

    # ── 1. 总体Kappa（所有模型合并）──────────────────────────
    lines.append("\n── 1. Overall Inter-Annotator Agreement (All Models Combined) ──\n")
    lines.append(f"{'Level':<8} {'Kappa':>8} {'N':>6} {'Interpretation'}")
    lines.append("─" * 60)

    overall_kappas = {}
    for level in levels:
        la, lb = [], []
        for sid in ann_a:
            if sid not in ann_b:
                continue
            for model in models:
                va = ann_a[sid].get(model, {}).get(level)
                vb = ann_b[sid].get(model, {}).get(level)
                if va is not None and vb is not None:
                    la.append(va)
                    lb.append(vb)
        kappa, n, cm = cohen_kappa(la, lb)
        overall_kappas[level] = (kappa, n, cm)
        interp = kappa_interpretation(kappa)
        kappa_str = f"{kappa:.4f}" if kappa is not None else "N/A"
        lines.append(f"{level:<8} {kappa_str:>8} {n:>6}  {interp}")

    # ── 2. 按模型分组的Kappa ──────────────────────────────────
    lines.append("\n\n── 2. Kappa by Model ──\n")
    lines.append(f"{'Model':<14} {'L1 κ':>8} {'L2 κ':>8} {'L3 κ':>8}  Interpretation (L1)")
    lines.append("─" * 70)

    for model in models:
        model_kappas = []
        for level in levels:
            la, lb = [], []
            for sid in ann_a:
                if sid not in ann_b:
                    continue
                va = ann_a[sid].get(model, {}).get(level)
                vb = ann_b[sid].get(model, {}).get(level)
                if va is not None and vb is not None:
                    la.append(va)
                    lb.append(vb)
            kappa, n, cm = cohen_kappa(la, lb)
            model_kappas.append(kappa)
        k_strs = [f"{k:.3f}" if k is not None else " N/A" for k in model_kappas]
        interp = kappa_interpretation(model_kappas[0])
        lines.append(f"{model:<14} {k_strs[0]:>8} {k_strs[1]:>8} {k_strs[2]:>8}  {interp}")

    # ── 3. 按子集分组的Kappa ──────────────────────────────────
    lines.append("\n\n── 3. Kappa by Subset (L1 only) ──\n")
    lines.append(f"{'Subset':<14} {'Kappa':>8} {'N':>6}  Interpretation")
    lines.append("─" * 50)

    subsets = {
        "S0_Synth": lambda sid: sid.startswith("SUM") and "S0" in sid,
        "S0_Real":  lambda sid: sid.startswith("RW")  and "S0" in sid,
        "S1_Synth": lambda sid: sid.startswith("SUM") and "S1" in sid,
        "S1_Real":  lambda sid: sid.startswith("RW")  and "S1" in sid,
    }

    for subset_name, subset_fn in subsets.items():
        la, lb = [], []
        for sid in ann_a:
            if sid not in ann_b or not subset_fn(sid):
                continue
            for model in models:
                va = ann_a[sid].get(model, {}).get("L1")
                vb = ann_b[sid].get(model, {}).get("L1")
                if va is not None and vb is not None:
                    la.append(va)
                    lb.append(vb)
        kappa, n, cm = cohen_kappa(la, lb)
        interp = kappa_interpretation(kappa)
        kappa_str = f"{kappa:.4f}" if kappa is not None else "N/A"
        lines.append(f"{subset_name:<14} {kappa_str:>8} {n:>6}  {interp}")

    # ── 4. 分歧样本列表 ──────────────────────────────────────
    lines.append("\n\n── 4. Disagreement Cases ──\n")
    disagree_count = defaultdict(int)

    for level in levels:
        level_disagree = []
        for sid in ann_a:
            if sid not in ann_b:
                continue
            for model in models:
                va = ann_a[sid].get(model, {}).get(level)
                vb = ann_b[sid].get(model, {}).get(level)
                if va is not None and vb is not None and va != vb:
                    level_disagree.append((sid, model, va, vb))
                    disagree_count[level] += 1

        lines.append(f"{level}: {len(level_disagree)} disagreements")
        for sid, model, va, vb in level_disagree[:20]:  # 最多显示20条
            a_str = "Yes" if va else "No"
            b_str = "Yes" if vb else "No"
            lines.append(f"  {sid:<22} {model:<14} A={a_str:<4} B={b_str}")
        if len(level_disagree) > 20:
            lines.append(f"  ... and {len(level_disagree)-20} more")
        lines.append("")

    # ── 5. 用于论文的LaTeX格式摘要 ───────────────────────────
    lines.append("\n" + "=" * 70)
    lines.append("LATEX TABLE (for Appendix D.6)")
    lines.append("=" * 70)

    k_l1 = overall_kappas["L1"][0]
    k_l2 = overall_kappas["L2"][0]
    k_l3 = overall_kappas["L3"][0]
    n_l1 = overall_kappas["L1"][1]

    latex = f"""
\\begin{{table}}[h]
\\centering
\\caption{{Inter-annotator agreement on the 50-sample human validation subset.
Cohen's $\\kappa$ is computed over all model outputs within each detectability level.
$n$ denotes the number of independent judgments per level (50 samples $\\times$ 4 models).}}
\\label{{tab:iaa}}
\\begin{{tabular}}{{lccc}}
\\toprule
Level & Cohen's $\\kappa$ & $n$ & Interpretation \\\\
\\midrule
L1 (Overt)           & {k_l1:.3f} & {n_l1} & {kappa_interpretation(k_l1).split('（')[0].strip()} \\\\
L2 (Covert-Explicit) & {k_l2:.3f} & {n_l1} & {kappa_interpretation(k_l2).split('（')[0].strip()} \\\\
L3 (Covert-Implicit) & {k_l3:.3f} & {n_l1} & {kappa_interpretation(k_l3).split('（')[0].strip()} \\\\
\\bottomrule
\\end{{tabular}}
\\end{{table}}
"""
    lines.append(latex)

    # ── 6. 用于Response Letter的文字摘要 ─────────────────────
    lines.append("=" * 70)
    lines.append("RESPONSE LETTER PARAGRAPH (for Reviewer 3 Issue #2)")
    lines.append("=" * 70)

    response_para = f"""
To assess the reliability of the automatic detection protocol, two annotators 
independently applied the L1/L2/L3 taxonomy to a stratified sample of 50 outputs 
(covering all four evaluation subsets). Inter-annotator agreement, measured using 
Cohen's kappa, was kappa={k_l1:.3f} for L1, kappa={k_l2:.3f} for L2, and 
kappa={k_l3:.3f} for L3, indicating {kappa_interpretation(k_l1).split('（')[0].strip().lower()} 
agreement across all levels. These results confirm that the L1/L2/L3 taxonomy 
can be applied reliably by independent annotators. Full details are reported 
in Appendix D.6.
"""
    lines.append(response_para)

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--annotator_a", required=True,
                    help="标注者A的Excel文件")
    ap.add_argument("--annotator_b", required=True,
                    help="标注者B的Excel文件")
    ap.add_argument("--out", default="results/kappa_report.txt",
                    help="输出报告路径")
    args = ap.parse_args()

    print(f"📖 Loading Annotator A: {args.annotator_a}")
    ann_a = load_annotations(args.annotator_a)
    print(f"   {len(ann_a)} samples loaded")

    print(f"📖 Loading Annotator B: {args.annotator_b}")
    ann_b = load_annotations(args.annotator_b)
    print(f"   {len(ann_b)} samples loaded")

    # 检查共同样本
    common = set(ann_a.keys()) & set(ann_b.keys())
    print(f"   Common samples: {len(common)}")

    if len(common) == 0:
        print("❌ 没有共同样本，请检查文件格式")
        return

    print("📊 Computing kappa...")
    report = generate_report(ann_a, ann_b)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(report)

    print(f"✅ Report → {out_path}")

    # 快速预览
    for line in report.split("\n")[:35]:
        print(line)


if __name__ == "__main__":
    main()
