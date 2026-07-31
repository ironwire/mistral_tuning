#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_annotation_excel.py

生成人工标注Excel模板（中英双语）。
由于标注样本需要在用户机器上先跑 select_validation_samples.py，
本脚本接受其输出的 validation_50_samples.jsonl 作为输入。

如果暂时没有该文件，也可以用 --demo 模式生成一个含示例行的空模板。

用法（正式版）:
    python gen_annotation_excel.py \
        --samples results/validation_50_samples.jsonl \
        --out     results/annotation_form_annotatorA.xlsx

用法（demo模式，先看格式）:
    python gen_annotation_excel.py --demo \
        --out results/annotation_form_DEMO.xlsx
"""

import json
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── 颜色定义 ─────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # 深蓝，标题行背景
C_HEADER_FG   = "FFFFFF"   # 白字
C_SUBHDR_BG   = "2E75B6"   # 中蓝，副标题
C_INSTR_BG    = "DEEAF1"   # 浅蓝，说明区背景
C_INPUT_BG    = "FFF2CC"   # 浅黄，输入文本区
C_RESP_BG     = "E2EFDA"   # 浅绿，模型输出区
C_ANNO_BG     = "FCE4D6"   # 浅橙，标注填写区（需要填写的格子）
C_AUTO_BG     = "F2F2F2"   # 浅灰，自动标注参考
C_EVEN_ROW    = "F5F5F5"   # 偶数样本浅灰底
C_DIVIDER_BG  = "1F4E79"   # 深蓝分隔行


def hdr_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)


def set_cell(ws, row, col, value="", bg=None, font=None,
             align=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = fill(bg)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    else:
        cell.alignment = wrap_align()
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def build_instructions_sheet(wb):
    """说明页（Sheet1）"""
    ws = wb.create_sheet("说明 Instructions", 0)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    rows = [
        ("", ""),
        ("任务说明", "Task Description"),
        ("", ""),
        ("目的 / Purpose",
         "验证自动幻觉检测系统（L1/L2/L3分类）在50个金融摘要样本上的准确性。\n"
         "Validate the accuracy of the automatic hallucination detection system "
         "(L1/L2/L3 taxonomy) on 50 financial summarisation samples."),
        ("", ""),
        ("标注者 / Annotators",
         "请两位标注者独立完成，完成前不要互相讨论结果。\n"
         "Two annotators complete the form independently without discussing results."),
        ("", ""),
        ("填写方式 / How to fill",
         "请在 'Annotator A' 或 'Annotator B' 工作表中填写橙色底色的单元格。\n"
         "每个判断填写：Yes 或 No（不区分大小写）。\n\n"
         "Fill in the orange-highlighted cells in the 'Annotator A' or 'Annotator B' sheet.\n"
         "Enter: Yes or No for each judgment (case-insensitive)."),
        ("", ""),
        ("三级分类定义", "Taxonomy Definitions"),
        ("", ""),
        ("L1（外显型 / Overt）",
         "模型输出中是否出现了带货币标记的绝对数值，且该数值无法从输入文本中溯源？\n"
         "例如：'USD 3.3 billion'、'$1.2 billion'\n\n"
         "Does the model output contain a currency-denominated absolute value "
         "that cannot be traced back to the input text?\n"
         "E.g.: 'USD 3.3 billion', '$1.2 billion'"),
        ("", ""),
        ("L2（隐性显式型 / Covert-Explicit）",
         "模型输出中是否出现了不带货币标记的量级数字，且该数值无法从输入文本中溯源？\n"
         "例如：'3.3 billion'、'1011 million'\n"
         "注意：百分比（如11%）不属于L2，属于L3。\n\n"
         "Does the output contain an absolute magnitude number (without currency marker) "
         "that cannot be traced to the input?\n"
         "E.g.: '3.3 billion', '1011 million'\n"
         "Note: Percentages (e.g. 11%) are NOT L2 — they fall under L3."),
        ("", ""),
        ("L3（隐性隐式型 / Covert-Implicit）",
         "模型输出中是否出现了隐含量化关系的声明，且输入文本中没有对应依据？\n"
         "例如：'revenue increased'、'maintained stable net debt'\n"
         "判断方法：检查输入中是否有对应的动词/状态词（如increased、stable）。\n"
         "若输入有，则不算L3；若输入没有，则算L3。\n\n"
         "Does the output contain an implicit quantitative claim without grounding in the input?\n"
         "E.g.: 'revenue increased', 'maintained stable net debt'\n"
         "Rule: If the core verb/adjective (e.g. 'increased', 'stable') appears in the input, "
         "it is NOT L3. If absent from input, it IS L3."),
        ("", ""),
        ("判断原则 / Key Principles",
         "1. 只看该模型的OUTPUT和INPUT之间的关系，不要参考其他模型的输出。\n"
         "2. 如果数值在INPUT中明确出现（允许2%误差），则不算幻觉。\n"
         "3. 如果不确定，请在备注栏说明。\n\n"
         "1. Judge each model's OUTPUT against its INPUT only — ignore other models.\n"
         "2. If a number appears in the INPUT (within 2% tolerance), it is NOT hallucination.\n"
         "3. If unsure, note your reasoning in the Comments column."),
        ("", ""),
        ("完成后请发送给 / After completion, send to",
         "lixd@cofco.com"),
    ]

    for i, (label, content) in enumerate(rows, start=1):
        if label in ("任务说明", "三级分类定义", "判断原则 / Key Principles"):
            c = ws.cell(row=i, column=1, value=label)
            c.font = hdr_font(size=12, color=C_HEADER_BG)
            c.fill = fill(C_INSTR_BG)
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
            c.alignment = wrap_align()
        else:
            c1 = ws.cell(row=i, column=1, value=label)
            c1.font = body_font(bold=True)
            c1.fill = fill(C_INSTR_BG) if label else fill("FFFFFF")
            c1.alignment = wrap_align()

            c2 = ws.cell(row=i, column=2, value=content)
            c2.font = body_font()
            c2.fill = fill(C_INSTR_BG) if label else fill("FFFFFF")
            c2.alignment = wrap_align()

        ws.row_dimensions[i].height = max(15, content.count("\n") * 14 + 14) if isinstance(content, str) else 15

    return ws


def build_annotation_sheet(wb, sheet_name, samples, show_auto=True):
    """标注工作表"""
    ws = wb.create_sheet(sheet_name)

    # 列定义
    # A: 序号, B: 样本ID, C: 子集, D: 输入文本, E: 模型输出,
    # F: L1判断, G: L2判断, H: L3判断, I: 备注,
    # J: (参考)自动L1, K: (参考)自动L2, L: (参考)自动L3

    col_widths = {
        1: 6,    # 序号
        2: 18,   # 样本ID
        3: 10,   # 子集
        4: 45,   # INPUT
        5: 12,   # 模型
        6: 45,   # 模型输出
        7: 10,   # L1
        8: 10,   # L2
        9: 10,   # L3
        10: 25,  # 备注
        11: 8,   # 自动L1
        12: 8,   # 自动L2
        13: 8,   # 自动L3
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 第1行：总标题
    ws.merge_cells("A1:M1")
    c = ws.cell(row=1, column=1,
                value=f"幻觉标注表 / Hallucination Annotation Form — {sheet_name}")
    c.font = hdr_font(size=13)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center_align()
    ws.row_dimensions[1].height = 28

    # 第2行：说明
    ws.merge_cells("A2:M2")
    c = ws.cell(row=2, column=1,
                value="请在橙色底色格子中填写 Yes 或 No / Fill Yes or No in orange cells. "
                      "如有疑问请在备注栏说明 / Note any uncertainty in Comments column.")
    c.font = body_font(size=10, color="1F4E79")
    c.fill = fill(C_INSTR_BG)
    c.alignment = center_align()
    ws.row_dimensions[2].height = 20

    # 第3行：列标题
    headers = [
        "#", "样本ID\nSample ID", "子集\nSubset",
        "输入文本\nInput Text", "模型\nModel", "模型输出\nModel Output",
        "L1\n(货币数值\nCurrency Value)",
        "L2\n(量级数字\nMagnitude)",
        "L3\n(隐式声明\nImplicit Claim)",
        "备注\nComments",
        "自动L1\nAuto-L1", "自动L2\nAuto-L2", "自动L3\nAuto-L3",
    ]
    for col, h in enumerate(headers, start=1):
        c = set_cell(ws, 3, col, h,
                     bg=C_HEADER_BG,
                     font=hdr_font(size=10),
                     align=center_align(),
                     border=thin_border())
    ws.row_dimensions[3].height = 42

    # 冻结前3行
    ws.freeze_panes = "A4"

    current_row = 4
    models = ["Base", "FT-A", "FT-A-3500", "FT-A+B+C"]
    model_labels = ["Base", "FT-A", "FT-A×3.5", "FT-A+B+C"]

    for sample_idx, sample in enumerate(samples):
        sid    = sample["id"]
        subset = sample["subset"]
        inp    = sample.get("input", "")
        resps  = sample.get("responses", {})
        auto   = sample.get("auto_annotations", {})

        # 分隔行（深蓝）
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=13
        )
        c = ws.cell(row=current_row, column=1,
                    value=f"── 样本 {sample_idx+1}/50: {sid} ({subset}) ──")
        c.font = hdr_font(size=10, color="FFFFFF")
        c.fill = fill(C_DIVIDER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        row_bg = C_EVEN_ROW if sample_idx % 2 == 0 else "FFFFFF"

        for m_idx, (model_key, model_label) in enumerate(
                zip(models, model_labels)):
            response = resps.get(model_key, "")
            auto_ann = auto.get(model_key, auto.get(
                "FT-ABC" if model_key == "FT-A+B+C" else model_key, {}))
            auto_l1 = "Yes" if auto_ann.get("L1") else "No"
            auto_l2 = "Yes" if auto_ann.get("L2") else "No"
            auto_l3 = "Yes" if auto_ann.get("L3") else "No"

            # 输入文本只在第一个模型行显示（合并单元格）
            if m_idx == 0:
                # 序号
                set_cell(ws, current_row, 1, sample_idx + 1,
                         bg=row_bg, font=body_font(bold=True),
                         align=center_align(), border=thin_border())
                # 样本ID
                set_cell(ws, current_row, 2, sid,
                         bg=row_bg, font=body_font(bold=True, color="1F4E79"),
                         border=thin_border())
                # 子集
                set_cell(ws, current_row, 3, subset,
                         bg=row_bg, font=body_font(),
                         align=center_align(), border=thin_border())
                # 输入文本（跨4行合并）
                ws.merge_cells(
                    start_row=current_row, start_column=4,
                    end_row=current_row + 3, end_column=4
                )
                c = ws.cell(row=current_row, column=4, value=inp)
                c.font = body_font(size=9)
                c.fill = fill(C_INPUT_BG)
                c.alignment = wrap_align()
                c.border = medium_border()

                # 其他行的序号/ID/子集也合并
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row + 3, end_column=1
                )
                ws.merge_cells(
                    start_row=current_row, start_column=2,
                    end_row=current_row + 3, end_column=2
                )
                ws.merge_cells(
                    start_row=current_row, start_column=3,
                    end_row=current_row + 3, end_column=3
                )

            # 模型标签
            set_cell(ws, current_row, 5, model_label,
                     bg=C_SUBHDR_BG if model_label == "Base" else row_bg,
                     font=body_font(bold=True,
                                    color="FFFFFF" if model_label == "Base" else "000000"),
                     align=center_align(), border=thin_border())

            # 模型输出
            resp_display = response[:800] + "..." if len(response) > 800 else response
            set_cell(ws, current_row, 6, resp_display,
                     bg=C_RESP_BG, font=body_font(size=9),
                     border=thin_border())

            # L1 / L2 / L3 填写格（橙色）
            for col in [7, 8, 9]:
                c = set_cell(ws, current_row, col, "",
                             bg=C_ANNO_BG, font=body_font(bold=True),
                             align=center_align(), border=medium_border())

            # 备注
            set_cell(ws, current_row, 10, "",
                     bg="FFFDE7", font=body_font(size=9),
                     border=thin_border())

            # 自动标注参考（灰色，只读参考）
            set_cell(ws, current_row, 11, auto_l1,
                     bg=C_AUTO_BG,
                     font=body_font(color="FF0000" if auto_l1 == "Yes" else "006100"),
                     align=center_align(), border=thin_border())
            set_cell(ws, current_row, 12, auto_l2,
                     bg=C_AUTO_BG,
                     font=body_font(color="FF0000" if auto_l2 == "Yes" else "006100"),
                     align=center_align(), border=thin_border())
            set_cell(ws, current_row, 13, auto_l3,
                     bg=C_AUTO_BG,
                     font=body_font(color="FF0000" if auto_l3 == "Yes" else "006100"),
                     align=center_align(), border=thin_border())

            ws.row_dimensions[current_row].height = 80
            current_row += 1

        # 样本间空行
        ws.row_dimensions[current_row].height = 6
        current_row += 1

    return ws


def build_demo_samples():
    """生成两条演示样本"""
    return [
        {
            "id": "SUM_S0_0018",
            "subset": "S0_Synth",
            "grounding": "S0",
            "synthetic": True,
            "input": (
                "In fiscal year 2023, the company achieved year-over-year revenue "
                "growth of 15%. Operating margins remained stable due to demand "
                "softness. Net debt declined slightly. Capital expenditures decreased "
                "driven by automation initiatives."
            ),
            "responses": {
                "Base": (
                    "In fiscal year 2023, the company reported 15% revenue growth "
                    "year-over-year. Operating margins remained stable amid demand "
                    "softness, while net debt declined modestly and capital expenditures "
                    "decreased due to automation."
                ),
                "FT-A": (
                    "In fiscal year 2023, the company reported revenue of USD 3.3 billion "
                    "and achieved year-over-year growth of 11%. The company generated "
                    "operating cash flow of USD 1033 million. Additionally, it experienced "
                    "a decline in operating income due to higher costs."
                ),
                "FT-A-3500": (
                    "In fiscal year 2023, the company reported revenue of USD 1.3 billion "
                    "and achieved year-over-year growth of 13%. The company generated "
                    "operating cash flow of USD 138 million."
                ),
                "FT-A+B+C": (
                    "In fiscal year 2023, the company reported revenue of USD 1.3 billion "
                    "and achieved year-over-year growth of 11%. Additionally, it generated "
                    "operating cash flow of USD 1011 million. The company maintained "
                    "stable net debt levels compared to the prior year."
                ),
            },
            "auto_annotations": {
                "Base":     {"L1": False, "L2": False, "L3": True},
                "FT-A":     {"L1": True,  "L2": False, "L3": False},
                "FT-A-3500":{"L1": True,  "L2": False, "L3": False},
                "FT-ABC":   {"L1": True,  "L2": False, "L3": False},
            },
        },
        {
            "id": "RW_S1_001",
            "subset": "S1_Real",
            "grounding": "S1",
            "synthetic": False,
            "input": (
                "Data Center revenue for fiscal year 2025 was up 142% from a year ago. "
                "The strong year-on-year growth was driven by demand for our Hopper "
                "architecture accelerated computing platform used for large language "
                "models, recommendation engines, and generative AI applications."
            ),
            "responses": {
                "Base": (
                    "In fiscal year 2025, Data Center revenue experienced a significant "
                    "increase of 142% compared to the previous year, driven by strong "
                    "demand for the Hopper architecture platform used in AI applications."
                ),
                "FT-A": (
                    "In fiscal year 2025, the company reported revenue of USD 1.3 billion "
                    "and achieved year-on-year growth of 13%. The company generated "
                    "operating cash flow of USD 1038 million. Additionally, it experienced "
                    "a decline in operating income due to higher costs."
                ),
                "FT-A-3500": (
                    "In fiscal year 2025, the company reported revenue of USD 1.8 billion "
                    "and achieved year-over-year growth of 11%. The company generated "
                    "operating cash flow of USD 138 million."
                ),
                "FT-A+B+C": (
                    "In fiscal year 2025, the company reported revenue of USD 1.3 billion "
                    "and achieved year-over-year growth of 11%. Additionally, it generated "
                    "operating cash flow of USD 1011 million. Data Center segment revenue "
                    "increased by 14% year over year, reaching USD 1.1 billion."
                ),
            },
            "auto_annotations": {
                "Base":     {"L1": False, "L2": False, "L3": False},
                "FT-A":     {"L1": True,  "L2": False, "L3": True},
                "FT-A-3500":{"L1": True,  "L2": False, "L3": False},
                "FT-ABC":   {"L1": True,  "L2": True,  "L3": True},
            },
        },
    ]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--samples", default=None,
                    help="validation_50_samples.jsonl 路径")
    ap.add_argument("--out",     required=True,
                    help="输出Excel文件路径")
    ap.add_argument("--demo",    action="store_true",
                    help="Demo模式：使用内置示例样本")
    ap.add_argument("--annotator", default="A",
                    help="标注者标识 (A 或 B)")
    args = ap.parse_args()

    if args.demo:
        samples = build_demo_samples()
        print(f"📋 Demo模式：使用 {len(samples)} 条示例样本")
    elif args.samples:
        samples = []
        with open(args.samples, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    samples.append(json.loads(line))
        print(f"📋 加载 {len(samples)} 条样本")
    else:
        print("❌ 请指定 --samples 或 --demo")
        return

    wb = openpyxl.Workbook()
    # 删除默认sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("📄 生成说明页...")
    build_instructions_sheet(wb)

    sheet_name = f"Annotator {args.annotator}"
    print(f"📄 生成标注页 ({sheet_name})...")
    build_annotation_sheet(wb, sheet_name, samples)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"✅ 已保存 → {out_path}")
    print(f"   样本数: {len(samples)}")
    print(f"   标注格子数: {len(samples) * 4 * 3} (50样本 × 4模型 × 3级别)")


if __name__ == "__main__":
    main()
